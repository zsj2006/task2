import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter
import os
from paddleocr import PaddleOCR
import numpy as np
import fitz  # PyMuPDF
import re
import json

# -------------------------- 配置参数 --------------------------
PDF_PATH = "../task1/345.pdf"
OUTPUT_EXCEL_PATH = "345_增强识别结果.xlsx"
DEBUG_TEXT_FILE = "ocr_raw_text_debug.txt"  # 保存原始OCR文本用于调试

# -------------------------- 工具函数 --------------------------
def pdf_to_ocr_with_debug(pdf_path):
    """将PDF转换为图片并进行OCR识别，保存原始文本用于调试"""
    print(f"正在OCR识别并保存调试信息...")

    pdf_doc = fitz.open(pdf_path)
    print(f"  PDF共 {len(pdf_doc)} 页")

    ocr = PaddleOCR(lang='ch')

    all_pages_text = []
    all_pages_debug = []

    for idx, page in enumerate(pdf_doc):
        print(f"  正在处理第 {idx + 1}/{len(pdf_doc)} 页...")

        # 渲染为高清晰度图片
        mat = fitz.Matrix(3, 3)  # 提高到3倍缩放
        pix = page.get_pixmap(matrix=mat)

        # 转换为numpy数组
        img_data = pix.tobytes("png")
        import io
        from PIL import Image
        img = Image.open(io.BytesIO(img_data))
        img_array = np.array(img)

        # OCR识别
        result = ocr.predict(img_array)

        if result and len(result) > 0:
            texts = result[0].get('rec_texts', [])
            scores = result[0].get('rec_scores', [])

            # 只保留置信度大于0.5的文本
            filtered_texts = []
            for text, score in zip(texts, scores):
                if text and score > 0.3:  # 降低阈值以获取更多信息
                    filtered_texts.append(f"{text} (置信度:{score:.2f})")

            page_text = " | ".join([t for t in texts if t])
            debug_text = "\n".join(filtered_texts)

            all_pages_text.append(page_text)
            all_pages_debug.append(f"===== 第{idx+1}页 =====\n{debug_text}\n")

            print(f"    识别到 {len(texts)} 个文本片段")
        else:
            all_pages_text.append("")
            all_pages_debug.append(f"===== 第{idx+1}页 =====\n识别失败\n")
            print(f"    识别失败")

    pdf_doc.close()

    # 保存调试文本
    with open(DEBUG_TEXT_FILE, 'w', encoding='utf-8') as f:
        f.write("\n".join(all_pages_debug))
    print(f"\n[调试] 原始OCR文本已保存到: {DEBUG_TEXT_FILE}")

    return all_pages_text

def extract_with_multiple_patterns(text, patterns):
    """使用多个模式尝试提取，返回第一个成功匹配的结果"""
    for pattern in patterns:
        match = re.search(pattern, text)
        if match:
            return match.group(1).strip()
    return ""

def parse_invoice_text_enhanced(text, page_num):
    """增强版发票信息解析"""
    parsed_data = {
        "票据序号": page_num + 1,
        "票据类型": "",
        "发票代码": "",
        "发票号码": "",
        "开票日期": "",
        "购买方名称": "",
        "购买方统一信用代码": "",
        "销售方名称": "",
        "销售方统一信用代码": "",
        "项目名称": "",
        "金额（不含税）": 0.0,
        "税率(%)": 0.0,
        "税额": 0.0,
        "价税合计": 0.0,
        "备注": ""
    }

    # 1. 发票类型（更灵活的匹配）
    if "增值税专用发票" in text:
        parsed_data["票据类型"] = "增值税专用发票"
    elif "增值税普通发票" in text:
        parsed_data["票据类型"] = "增值税普通发票"
    elif "电子发票" in text or "电子" in text:
        parsed_data["票据类型"] = "电子发票"
    elif "通用机打发票" in text:
        parsed_data["票据类型"] = "通用机打发票"

    # 2. 发票代码和号码
    code_patterns = [
        r'发票代码[：:]\s*(\d+)',
        r'代码[：:]\s*(\d{12})',
    ]
    parsed_data["发票代码"] = extract_with_multiple_patterns(text, code_patterns)

    no_patterns = [
        r'发票号码[：:]\s*(\d+)',
        r'号码[：:]\s*(\d{8})',
    ]
    parsed_data["发票号码"] = extract_with_multiple_patterns(text, no_patterns)

    # 3. 开票日期
    date_patterns = [
        r'开票日期[：:]\s*(\d{4}年\d{1,2}月\d{1,2}日)',
        r'开票日期[：:]\s*(\d{4}-\d{1,2}-\d{1,2})',
        r'(\d{4}年\d{1,2}月\d{1,2}日)',
    ]
    parsed_data["开票日期"] = extract_with_multiple_patterns(text, date_patterns)

    # 4. 购买方名称（更全面的模式）
    buyer_patterns = [
        r'购买方[\s\S]{0,50}名称[：:]\s*([^\n]+?)(?=\s+纳税人识别号|统一社会信用代码|密区码|$)',
        r'购[\s\S]{0,30}名称[：:]\s*([^\n]+?)(?=\s+纳税人识别号|统一社会信用代码|$)',
        r'客户名称[：:]\s*([^\n]+)',
        r'抬头[：:]\s*([^\n]+)',
        r'名称[：:]\s*([^\n]{2,20}?)(?=\s+纳税人识别号)',
    ]
    parsed_data["购买方名称"] = extract_with_multiple_patterns(text, buyer_patterns)

    # 5. 销售方名称
    seller_patterns = [
        r'销售方[\s\S]{0,50}名称[：:]\s*([^\n]+?)(?=\s+纳税人识别号|统一社会信用代码|备注|$)',
        r'销[\s\S]{0,30}名称[：:]\s*([^\n]+?)(?=\s+纳税人识别号|统一社会信用代码|$)',
        r'销售商[：:]\s*([^\n]+)',
        r'商家名称[：:]\s*([^\n]+)',
    ]
    parsed_data["销售方名称"] = extract_with_multiple_patterns(text, seller_patterns)

    # 6. 购买方税号
    buyer_tax_patterns = [
        r'购买方[\s\S]{0,100}纳税人识别号[：:]\s*([A-Z0-9]{18,20})',
        r'购买方[\s\S]{0,100}统一社会信用代码[：:]\s*([A-Z0-9]{18,20})',
        r'客户[\s\S]{0,50}税号[：:]\s*([A-Z0-9]{18,20})',
        r'纳税人识别号[：:]\s*([A-Z0-9]{18,20})',
    ]
    parsed_data["购买方统一信用代码"] = extract_with_multiple_patterns(text, buyer_tax_patterns)

    # 7. 销售方税号
    seller_tax_patterns = [
        r'销售方[\s\S]{0,100}纳税人识别号[：:]\s*([A-Z0-9]{18,20})',
        r'销售方[\s\S]{0,100}统一社会信用代码[：:]\s*([A-Z0-9]{18,20})',
        r'销售商[\s\S]{0,50}税号[：:]\s*([A-Z0-9]{18,20})',
    ]
    parsed_data["销售方统一信用代码"] = extract_with_multiple_patterns(text, seller_tax_patterns)

    # 8. 项目名称
    item_patterns = [
        r'\*[\u4e00-\u9fa5a-zA-Z0-9]+\*\s*([^\n]+?)(?=\s+规格型号|单位|数量|单价|$)',
        r'货物或应税劳务[\s\S]{0,20}名称[：:]\s*([^\n]+)',
        r'项目[：:]\s*([^\n]+)',
        r'商品名称[：:]\s*([^\n]+)',
        r'服务名称[：:]\s*([^\n]+)',
        r'名称[：:]\s*([^\n]{2,30}?)(?=\s+规格型号|单位)',
    ]
    parsed_data["项目名称"] = extract_with_multiple_patterns(text, item_patterns)

    # 9. 价税合计（最重要的字段）
    total_patterns = [
        r'价税合计[\s\S]{0,20}大写[）:]?\s*([^\n]+)',
        r'价税合计[\s\S]{0,30}[小写][）:]?\s*[￥¥]?\s*([\d,]+\.?\d*)',
        r'[（(]小写[）)]\s*[￥¥]?\s*([\d,]+\.?\d*)',
        r'合计[）:]?\s*[￥¥]?\s*([\d,]+\.?\d*)',
        r'总金额[）:]?\s*[￥¥]?\s*([\d,]+\.?\d*)',
    ]
    for pattern in total_patterns:
        match = re.search(pattern, text)
        if match:
            amount_str = match.group(1).replace('¥', '').replace('￥', '').replace(',', '').strip()
            # 过滤掉大写数字
            if not any(c in amount_str for c in '壹贰叁肆伍陆柒捌玖拾佰仟万亿元整角分'):
                try:
                    parsed_data["价税合计"] = float(amount_str)
                    break
                except:
                    pass

    # 10. 税率
    tax_rate_patterns = [
        r'税率[】】]?\s*(\d+)%',
        r'税率[：:]\s*(\d+)',
    ]
    for pattern in tax_rate_patterns:
        match = re.search(pattern, text)
        if match:
            try:
                parsed_data["税率(%)"] = float(match.group(1))
                break
            except:
                pass

    # 11. 税额
    tax_patterns = [
        r'税额[\s\S]{0,10}[）:]?\s*[￥¥]?\s*([\d,]+\.?\d*)',
        r'税额[）:]?\s*[￥¥]?\s*([\d,]+\.?\d*)',
    ]
    for pattern in tax_patterns:
        match = re.search(pattern, text)
        if match:
            tax_str = match.group(1).replace('¥', '').replace('￥', '').replace(',', '').strip()
            try:
                parsed_data["税额"] = float(tax_str)
                break
            except:
                pass

    # 12. 金额（不含税）
    amount_patterns = [
        r'金额[\s\S]{0,10}[不含税][）:]?\s*[￥¥]?\s*([\d,]+\.?\d*)',
        r'金额[）:]?\s*[￥¥]?\s*([\d,]+\.?\d*)',
        r'不含税金额[）:]?\s*[￥¥]?\s*([\d,]+\.?\d*)',
    ]
    for pattern in amount_patterns:
        match = re.search(pattern, text)
        if match:
            amount_str = match.group(1).replace('¥', '').replace('￥', '').replace(',', '').strip()
            try:
                parsed_data["金额（不含税）"] = float(amount_str)
                break
            except:
                pass

    # 计算逻辑：如果只有部分数据，尝试计算
    if parsed_data["价税合计"] > 0:
        if parsed_data["税额"] == 0 and parsed_data["税率(%)"] > 0:
            # 已知价税合计和税率，计算税额
            parsed_data["税额"] = round(parsed_data["价税合计"] / (1 + parsed_data["税率(%)"]/100) * (parsed_data["税率(%)"]/100), 2)
        if parsed_data["金额（不含税）"] == 0:
            # 已知价税合计和税额，计算不含税金额
            parsed_data["金额（不含税）"] = round(parsed_data["价税合计"] - parsed_data["税额"], 2)

    return parsed_data

def generate_excel_enhanced(parsed_data_list):
    """生成增强版Excel报表"""
    wb = Workbook()
    ws = wb.active
    ws.title = "票据识别结果"

    # 标题
    ws['A1'] = '票据识别结果表 - 增强OCR版'
    ws['A1'].font = Font(size=18, bold=True, color='1F4E78')
    ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
    ws.merge_cells('A1:N1')

    ws['A2'] = f'制表日期: {pd.Timestamp.now().strftime("%Y年%m月%d日")}  |  共识别 {len(parsed_data_list)} 页'
    ws['A2'].font = Font(size=11, color='0070C0')
    ws.merge_cells('A2:N2')

    # 表头
    headers = [
        "票据序号", "票据类型", "发票代码", "发票号码", "开票日期",
        "购买方名称", "购买方统一信用代码", "销售方名称", "销售方统一信用代码",
        "项目名称", "金额（不含税）", "税率(%)", "税额", "价税合计"
    ]

    header_row = 3
    header_font = Font(bold=True, color="FFFFFF", size=10)
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin")
    )

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=header_row, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = thin_border

    # 数据
    for row_idx, data in enumerate(parsed_data_list, header_row + 1):
        for col, key in enumerate(headers, 1):
            value = data.get(key, "")
            cell = ws.cell(row=row_idx, column=col, value=value)
            cell.alignment = Alignment(horizontal="left" if col > 4 else "center", vertical="center")
            cell.border = thin_border

    # 列宽
    column_widths = [8, 12, 15, 12, 15, 25, 20, 25, 20, 20, 12, 8, 12, 12]
    for col_idx, width in enumerate(column_widths, 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    ws.freeze_panes = 'A4'

    wb.save(OUTPUT_EXCEL_PATH)
    print(f"\n[OK] Excel已生成：{OUTPUT_EXCEL_PATH}")

    os.startfile(OUTPUT_EXCEL_PATH)

# -------------------------- 主流程 --------------------------
def main():
    print("=" * 70)
    print("增强版 OCR 票据识别系统")
    print("=" * 70)

    # OCR识别
    print(f"\n[1/2] OCR识别: {PDF_PATH}")
    all_pages_text = pdf_to_ocr_with_debug(PDF_PATH)

    # 解析信息
    print(f"\n[2/2] 解析票据信息...")
    parsed_data_list = []
    for idx, text in enumerate(all_pages_text):
        print(f"  解析第 {idx + 1} 页...")
        parsed_data = parse_invoice_text_enhanced(text, idx)
        parsed_data_list.append(parsed_data)

        # 显示关键信息
        print(f"    类型: {parsed_data['票据类型']}")
        print(f"    日期: {parsed_data['开票日期']}")
        print(f"    购买方: {parsed_data['购买方名称'][:30] if parsed_data['购买方名称'] else '未识别'}")
        print(f"    销售方: {parsed_data['销售方名称'][:30] if parsed_data['销售方名称'] else '未识别'}")
        print(f"    项目: {parsed_data['项目名称'][:30] if parsed_data['项目名称'] else '未识别'}")
        print(f"    金额: {parsed_data['价税合计']}")
        print()

    # 生成Excel
    print(f"生成Excel报表...")
    generate_excel_enhanced(parsed_data_list)

    print("\n处理完成！")
    print(f"调试文件: {DEBUG_TEXT_FILE}")
    print(f"结果文件: {OUTPUT_EXCEL_PATH}")

if __name__ == "__main__":
    main()

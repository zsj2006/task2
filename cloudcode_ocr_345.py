import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter
import os
from paddleocr import PaddleOCR
import numpy as np
import fitz  # PyMuPDF
import re

# -------------------------- 配置参数 --------------------------
# 345.pdf 路径
PDF_PATH = "../task1/345.pdf"
# 输出Excel路径
OUTPUT_EXCEL_PATH = "345_票据识别结果.xlsx"

# -------------------------- 工具函数 --------------------------
def pdf_to_ocr_images(pdf_path):
    """将PDF转换为图片并进行OCR识别"""
    print(f"正在将PDF转换为图片并OCR识别...")

    # 打开PDF
    pdf_doc = fitz.open(pdf_path)
    print(f"  PDF共 {len(pdf_doc)} 页")

    # 初始化OCR
    ocr = PaddleOCR(lang='ch')

    all_pages_text = []

    for idx, page in enumerate(pdf_doc):
        print(f"  正在处理第 {idx + 1}/{len(pdf_doc)} 页...")

        # 将PDF页面渲染为图片
        mat = fitz.Matrix(2, 2)  # 2倍缩放提高清晰度
        pix = page.get_pixmap(matrix=mat)

        # 转换为numpy数组
        img_data = pix.tobytes("png")
        import io
        from PIL import Image
        img = Image.open(io.BytesIO(img_data))
        img_array = np.array(img)

        # OCR识别
        result = ocr.predict(img_array)

        if result and len(result) > 0:
            texts = result[0].get('rec_texts', [])
            page_text = "\n".join([t for t in texts if t])
            all_pages_text.append(page_text)
            print(f"    识别到 {len(texts)} 个文本片段")
        else:
            all_pages_text.append("")
            print(f"    识别失败")

    pdf_doc.close()
    return all_pages_text

def parse_invoice_text(text, page_num):
    """解析发票文本信息"""
    parsed_data = {
        "票据序号": page_num + 1,
        "购买方名称": "",
        "购买方统一信用代码": "",
        "销售方名称": "",
        "销售方统一信用代码": "",
        "票据类型": "",
        "项目名称": "",
        "金额（不含税）": 0.0,
        "税率(%)": 0.0,
        "税额": 0.0,
        "价税合计": 0.0,
        "备注": ""
    }

    # 发票类型
    if "增值税专用发票" in text:
        parsed_data["票据类型"] = "增值税专用发票"
    elif "增值税普通发票" in text:
        parsed_data["票据类型"] = "增值税普通发票"
    elif "电子发票" in text:
        parsed_data["票据类型"] = "电子发票"
    elif "通用机打发票" in text:
        parsed_data["票据类型"] = "通用机打发票"

    # 购买方名称
    buyer_patterns = [
        r'购买方.*?名称[：:]\s*([^\s]+(?:\s+[^\s]+)*?)(?=\s+纳税人|统一社会信用代码|$)',
        r'购.*?名称[：:]\s*([^\s]+(?:\s+[^\s]+)*?)(?=\s|$)',
        r'客户名称[：:]\s*([^\s]+(?:\s+[^\s]+)*?)(?=\s|$)',
        r'抬头[：:]\s*([^\s]+(?:\s+[^\s]+)*?)(?=\s|$)',
    ]
    for pattern in buyer_patterns:
        match = re.search(pattern, text)
        if match:
            parsed_data["购买方名称"] = match.group(1).strip()
            break

    # 销售方名称
    seller_patterns = [
        r'销售方.*?名称[：:]\s*([^\s]+(?:\s+[^\s]+)*?)(?=\s+纳税人|统一社会信用代码|$)',
        r'销.*?名称[：:]\s*([^\s]+(?:\s+[^\s]+)*?)(?=\s|$)',
        r'销售商[：:]\s*([^\s]+(?:\s+[^\s]+)*?)(?=\s|$)',
        r'商家名称[：:]\s*([^\s]+(?:\s+[^\s]+)*?)(?=\s|$)',
    ]
    for pattern in seller_patterns:
        match = re.search(pattern, text)
        if match:
            parsed_data["销售方名称"] = match.group(1).strip()
            break

    # 购买方税号
    buyer_tax_patterns = [
        r'购买方.*?纳税人识别号[：:]\s*([A-Z0-9]+)',
        r'购买方.*?统一社会信用代码[：:]\s*([A-Z0-9]+)',
        r'客户.*?税号[：:]\s*([A-Z0-9]+)',
    ]
    for pattern in buyer_tax_patterns:
        match = re.search(pattern, text)
        if match:
            parsed_data["购买方统一信用代码"] = match.group(1)
            break

    # 销售方税号
    seller_tax_patterns = [
        r'销售方.*?纳税人识别号[：:]\s*([A-Z0-9]+)',
        r'销售方.*?统一社会信用代码[：:]\s*([A-Z0-9]+)',
        r'销售商.*?税号[：:]\s*([A-Z0-9]+)',
    ]
    for pattern in seller_tax_patterns:
        match = re.search(pattern, text)
        if match:
            parsed_data["销售方统一信用代码"] = match.group(1)
            break

    # 项目名称
    item_patterns = [
        r'\*[\u4e00-\u9fa5a-zA-Z]+\*\s*([^\d\n]+?)(?=\s+规格型号|单位|数量|单价|$)',
        r'货物或应税劳务.*?名称[：:]\s*([^\s]+)',
        r'项目[：:]\s*([^\s]+)',
        r'商品名称[：:]\s*([^\s]+)',
        r'服务名称[：:]\s*([^\s]+)',
    ]
    for pattern in item_patterns:
        match = re.search(pattern, text)
        if match:
            parsed_data["项目名称"] = match.group(1).strip()
            break

    # 价税合计
    total_patterns = [
        r'[价税合计][小写][）:]?\s*[￥¥]?\s*([\d,]+\.?\d*)',
        r'合计[）:]?\s*[￥¥]?\s*([\d,]+\.?\d*)',
        r'总金额[）:]?\s*[￥¥]?\s*([\d,]+\.?\d*)',
    ]
    for pattern in total_patterns:
        match = re.search(pattern, text)
        if match:
            amount_str = match.group(1).replace('¥', '').replace('￥', '').replace(',', '').strip()
            try:
                parsed_data["价税合计"] = float(amount_str)
                break
            except:
                pass

    # 税率
    tax_rate_match = re.search(r'税率[】】]?\s*(\d+)%', text)
    if tax_rate_match:
        parsed_data["税率(%)"] = float(tax_rate_match.group(1))

    # 税额
    tax_patterns = [
        r'[税额][）:]?\s*[￥¥]?\s*([\d,]+\.?\d*)',
        r'税额[）:]?\s*[￥¥]?\s*([\d,]+\.?\d*)',
    ]
    for pattern in tax_patterns:
        match = re.search(pattern, text)
        if match:
            tax_str = match.group(1).replace('¥', '').replace('￥', '').replace(',', '').strip()
            try:
                parsed_data["税额"] = float(tax_str)
                break
            except:
                pass

    # 金额（不含税）
    amount_patterns = [
        r'[金额][不含税][）:]?\s*[￥¥]?\s*([\d,]+\.?\d*)',
        r'金额[）:]?\s*[￥¥]?\s*([\d,]+\.?\d*)',
    ]
    for pattern in amount_patterns:
        match = re.search(pattern, text)
        if match:
            amount_str = match.group(1).replace('¥', '').replace('￥', '').replace(',', '').strip()
            try:
                parsed_data["金额（不含税）"] = float(amount_str)
                break
            except:
                pass

    # 如果没有金额但有价税合计和税额，计算金额
    if parsed_data["金额（不含税）"] == 0 and parsed_data["价税合计"] > 0 and parsed_data["税额"] > 0:
        parsed_data["金额（不含税）"] = round(parsed_data["价税合计"] - parsed_data["税额"], 2)

    # 判断票据类型（基于项目名称）
    if parsed_data["项目名称"]:
        item_lower = parsed_data["项目名称"].lower()
        if "运输" in item_lower or "客运" in item_lower or "滴滴" in item_lower:
            parsed_data["票据类型"] = "运输服务"
        elif "餐饮" in item_lower or "餐费" in item_lower:
            parsed_data["票据类型"] = "餐饮服务"
        elif "住宿" in item_lower or "酒店" in item_lower:
            parsed_data["票据类型"] = "住宿服务"
        elif "食品" in item_lower or "烘焙" in item_lower or "商品" in item_lower:
            parsed_data["票据类型"] = "商品采购"

    return parsed_data

def generate_excel(parsed_data_list):
    """生成标准化Excel报表（CloudCode风格）"""
    # 创建工作簿
    wb = Workbook()
    ws = wb.active
    ws.title = "票据识别结果"

    # 定义表头
    headers = [
        "票据序号", "票据类型", "购买方名称", "购买方统一信用代码",
        "销售方名称", "销售方统一信用代码", "项目名称", "金额（不含税）",
        "税率(%)", "税额", "价税合计", "备注"
    ]

    # 写入表头并设置样式
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    alignment = Alignment(horizontal="center", vertical="center")
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin")
    )

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = alignment
        cell.border = thin_border

    # 写入数据
    for row_idx, data in enumerate(parsed_data_list, 2):
        for col, key in enumerate(headers, 1):
            cell = ws.cell(row=row_idx, column=col, value=data.get(key, ""))
            cell.alignment = Alignment(horizontal="left" if col > 2 else "center", vertical="center")
            cell.border = thin_border

    # 自动调整列宽
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column].width = adjusted_width

    # 保存Excel
    wb.save(OUTPUT_EXCEL_PATH)
    print(f"\n[OK] Excel已生成：{OUTPUT_EXCEL_PATH}")

    # 自动打开
    os.startfile(OUTPUT_EXCEL_PATH)

# -------------------------- 主流程 --------------------------
def main():
    print("=" * 60)
    print("CloudCode OCR - 票据识别系统 (345.pdf)")
    print("=" * 60)

    # 步骤1: PDF转图片并OCR
    print(f"\n[1/2] 正在处理PDF: {PDF_PATH}")
    all_pages_text = pdf_to_ocr_images(PDF_PATH)

    # 步骤2: 解析每页信息
    print(f"\n[2/2] 正在解析票据信息...")
    parsed_data_list = []
    for idx, text in enumerate(all_pages_text):
        print(f"  解析第 {idx + 1} 页...")
        parsed_data = parse_invoice_text(text, idx)
        parsed_data_list.append(parsed_data)

        # 显示关键信息
        print(f"    类型: {parsed_data['票据类型']}")
        print(f"    项目: {parsed_data['项目名称'][:30] if parsed_data['项目名称'] else '未识别'}")
        print(f"    金额: {parsed_data['价税合计']}")

    # 生成Excel
    print(f"\n生成Excel报表...")
    generate_excel(parsed_data_list)

    print("\n处理完成！")

if __name__ == "__main__":
    main()

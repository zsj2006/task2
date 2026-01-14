import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter
import os
from paddleocr import PaddleOCR
import numpy as np
import fitz  # PyMuPDF
import re

# -------------------------- 配置参数 --------------------------
# 234.pdf 路径
PDF_PATH = "../task1/批量发票/234.pdf"
# 输出Excel路径
OUTPUT_EXCEL_PATH = "234_票据识别结果.xlsx"

# -------------------------- 工具函数 --------------------------
def pdf_to_ocr_images(pdf_path):
    """将PDF转换为图片并进行OCR识别"""
    print(f"正在将PDF转换为图片并OCR识别...")

    # 打开PDF
    pdf_doc = fitz.open(pdf_path)
    print(f"  PDF共 {len(pdf_doc)} 页")

    # 初始化OCR
    ocr = PaddleOCR(lang='ch')

    all_pages_text = []

    for idx, page in enumerate(pdf_doc):
        print(f"  正在处理第 {idx + 1}/{len(pdf_doc)} 页...")

        # 将PDF页面渲染为图片
        mat = fitz.Matrix(2, 2)  # 2倍缩放提高清晰度
        pix = page.get_pixmap(matrix=mat)

        # 转换为numpy数组
        img_data = pix.tobytes("png")
        import io
        from PIL import Image
        img = Image.open(io.BytesIO(img_data))
        img_array = np.array(img)

        # OCR识别
        result = ocr.predict(img_array)

        if result and len(result) > 0:
            texts = result[0].get('rec_texts', [])
            page_text = "\n".join([t for t in texts if t])
            all_pages_text.append(page_text)
            print(f"    识别到 {len(texts)} 个文本片段")
        else:
            all_pages_text.append("")
            print(f"    识别失败")

    pdf_doc.close()
    return all_pages_text

def parse_invoice_text(text):
    """解析发票文本信息"""
    info = {
        "票据序号": 1,
        "购买方名称": "",
        "购买方统一信用代码": "",
        "销售方名称": "",
        "销售方统一信用代码": "",
        "票据类型": "电子发票",
        "项目名称": "",
        "金额（不含税）": 0.0,
        "税率(%)": 0.0,
        "税额": 0.0,
        "价税合计": 0.0,
        "备注": ""
    }

    # 发票类型
    if "增值税专用发票" in text:
        info["票据类型"] = "增值税专用发票"
    elif "增值税普通发票" in text:
        info["票据类型"] = "增值税普通发票"
    elif "电子发票" in text:
        info["票据类型"] = "电子发票"

    # 购买方名称
    buyer_patterns = [
        r'购买方.*?名称[：:]\s*([^\s]+(?:\s+[^\s]+)*?)(?=\s+纳税人|统一社会信用代码)',
        r'购.*?名称[：:]\s*([^\s]+(?:\s+[^\s]+)*?)(?=\s|$)',
    ]
    for pattern in buyer_patterns:
        match = re.search(pattern, text)
        if match:
            info["购买方名称"] = match.group(1).strip()
            break

    # 销售方名称
    seller_patterns = [
        r'销售方.*?名称[：:]\s*([^\s]+(?:\s+[^\s]+)*?)(?=\s+纳税人|统一社会信用代码)',
        r'销.*?名称[：:]\s*([^\s]+(?:\s+[^\s]+)*?)(?=\s|$)',
    ]
    for pattern in seller_patterns:
        match = re.search(pattern, text)
        if match:
            info["销售方名称"] = match.group(1).strip()
            break

    # 购买方税号
    buyer_tax_patterns = [
        r'购买方.*?纳税人识别号[：:]\s*([A-Z0-9]+)',
        r'购买方.*?统一社会信用代码[：:]\s*([A-Z0-9]+)',
    ]
    for pattern in buyer_tax_patterns:
        match = re.search(pattern, text)
        if match:
            info["购买方统一信用代码"] = match.group(1)
            break

    # 销售方税号
    seller_tax_patterns = [
        r'销售方.*?纳税人识别号[：:]\s*([A-Z0-9]+)',
        r'销售方.*?统一社会信用代码[：:]\s*([A-Z0-9]+)',
    ]
    for pattern in seller_tax_patterns:
        match = re.search(pattern, text)
        if match:
            info["销售方统一信用代码"] = match.group(1)
            break

    # 项目名称（从 *税收分类* 项目名称 格式提取）
    item_patterns = [
        r'\*[\u4e00-\u9fa5a-zA-Z]+\*\s*([^\d\n]+?)(?=\s+规格型号|单位|数量|$)',
        r'货物或应税劳务.*?名称[：:]\s*([^\s]+)',
    ]
    for pattern in item_patterns:
        match = re.search(pattern, text)
        if match:
            info["项目名称"] = match.group(1).strip()
            break

    # 金额相关
    amount_match = re.search(r'[金额价税合计].*?([¥￥]?\s*[\d,]+\.?\d*)', text)
    if amount_match:
        amount_str = amount_match.group(1).replace('¥', '').replace('￥', '').replace(',', '').strip()
        try:
            info["价税合计"] = float(amount_str)
        except:
            pass

    # 税率
    tax_rate_match = re.search(r'税率[】】]?\s*(\d+)%', text)
    if tax_rate_match:
        info["税率(%)"] = float(tax_rate_match.group(1))

    # 税额
    tax_match = re.search(r'税额[】】]?\s*[¥￥]?\s*([\d,]+\.?\d*)', text)
    if tax_match:
        tax_str = tax_match.group(1).replace('¥', '').replace('￥', '').replace(',', '').strip()
        try:
            info["税额"] = float(tax_str)
        except:
            pass

    # 金额（不含税）
    if info["价税合计"] > 0 and info["税额"] > 0:
        info["金额（不含税）"] = round(info["价税合计"] - info["税额"], 2)

    return info

def generate_excel(all_pages_info):
    """生成标准化Excel报表"""
    # 创建工作簿
    wb = Workbook()
    ws = wb.active
    ws.title = "票据识别结果"

    # 插入标题行
    ws['A1'] = '票据识别结果表 - 234.pdf OCR识别'
    ws['A1'].font = Font(size=18, bold=True, color='1F4E78')
    ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
    ws.merge_cells(f'A1:L1')

    # 插入日期
    ws['A2'] = f'制表日期: {pd.Timestamp.now().strftime("%Y年%m月%d日")}  |  共识别 {len(all_pages_info)} 页'
    ws['A2'].font = Font(size=11, color='0070C0')
    ws.merge_cells(f'A2:L2')

    # 定义表头（第3行）
    headers = [
        "票据序号", "票据类型", "购买方名称", "购买方统一信用代码",
        "销售方名称", "销售方统一信用代码", "项目名称", "金额（不含税）",
        "税率(%)", "税额", "价税合计", "备注"
    ]

    # 写入表头并设置样式
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin")
    )

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=3, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = thin_border

    # 写入数据
    for row_idx, page_info in enumerate(all_pages_info, 4):
        for col, key in enumerate(headers, 1):
            cell = ws.cell(row=row_idx, column=col, value=page_info.get(key, ""))
            cell.alignment = Alignment(horizontal="left" if col > 2 else "center", vertical="center")
            cell.border = thin_border

    # 自动调整列宽
    for col in range(1, ws.max_column + 1):
        col_letter = get_column_letter(col)
        ws.column_dimensions[col_letter].width = 18

    # 冻结首行
    ws.freeze_panes = 'A4'

    # 保存Excel
    wb.save(OUTPUT_EXCEL_PATH)
    print(f"\n[OK] Excel已生成：{OUTPUT_EXCEL_PATH}")

    # 自动打开
    os.startfile(OUTPUT_EXCEL_PATH)

# -------------------------- 主流程 --------------------------
def main():
    print("=" * 60)
    print("CloudCode OCR - 票据识别系统 (234.pdf)")
    print("=" * 60)

    # 步骤1: PDF转图片并OCR
    print(f"\n[1/2] 正在处理PDF: {PDF_PATH}")
    all_pages_text = pdf_to_ocr_images(PDF_PATH)

    # 步骤2: 解析每页信息
    print(f"\n[2/2] 正在解析票据信息...")
    all_pages_info = []
    for idx, text in enumerate(all_pages_text):
        print(f"  解析第 {idx + 1} 页...")
        info = parse_invoice_text(text)
        info["票据序号"] = idx + 1
        all_pages_info.append(info)

        # 显示关键信息
        print(f"    项目: {info['项目名称'][:30] if info['项目名称'] else '未识别'}")
        print(f"    金额: {info['价税合计']}")

    # 生成Excel
    print(f"\n生成Excel报表...")
    generate_excel(all_pages_info)

    print("\n处理完成！")

if __name__ == "__main__":
    main()
